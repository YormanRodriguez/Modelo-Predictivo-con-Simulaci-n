# services/export_service.py - Servicio para exportar predicciones a Excel
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


@dataclass
class RegionalInfo:
    """Informacion de regional."""

    code: str
    nombre: str


@dataclass
class ExportConfig:
    """Configuración para la exportación de predicciones."""

    predictions_dict: dict[str, Any]
    regional_code: str
    regional_nombre: str
    output_dir: Path | None = None
    include_confidence_intervals: bool = True
    model_info: dict[str, Any] | None = None


class ExportService:
    """Servicio para exportar predicciones SAIDI a Excel con formato profesional."""

    def __init__(self):
        self.last_export_path = None

    def export_predictions_to_excel(self, config: ExportConfig):
        """
        Exportar predicciones a archivo Excel con formato.

        Args:
            config: Configuración de exportación con predicciones, códigos y opciones

        Returns:
            str: Ruta del archivo exportado o None si hay error

        """
        if not config.predictions_dict:
            raise ValueError("No hay predicciones para exportar")

        return self._export_with_config(config)

    def _export_with_config(self, config: ExportConfig) -> str | None:
        """Exporta usando configuración interna."""
        try:
            filepath = self._prepare_filepath(config)

            # Si el archivo ya existe, eliminarlo antes de crear uno nuevo
            if filepath.exists():
                self._remove_existing_file(filepath)

            df = self._create_dataframe(config)
            wb = self._create_workbook(df, config)

            if config.model_info:
                self._add_model_info_sheet(wb, config)

            # Guardar y verificar
            self._save_and_verify_workbook(wb, filepath)

            self.last_export_path = str(filepath)
            return str(filepath)

        except (OSError, ValueError) as e:
            print(f"Error exportando predicciones: {e!s}")
            return None

    def _save_and_verify_workbook(self, wb, filepath):
        """Guardar workbook y verificar que se creó correctamente."""
        wb.save(str(filepath))

        if not filepath.exists():
            msg = f"El archivo no se creó correctamente: {filepath}"
            raise OSError(msg)

    def _remove_existing_file(self, filepath):
        """Eliminar archivo existente con manejo de errores."""
        try:
            filepath.unlink()
        except (OSError, PermissionError) as e:
            print(f"Advertencia al eliminar archivo existente: {e}")

    def _prepare_filepath(self, config: ExportConfig) -> Path:
        """Prepara el directorio y nombre del archivo."""
        output_dir = config.output_dir or Path.home() / "Desktop"
        output_dir = Path(output_dir)

        if not output_dir.exists():
            output_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now(tz=UTC).strftime("%Y%m%d_%H%M%S")
        regional_clean = config.regional_nombre.replace(" ", "_")
        filename = f"Predicciones_SAIDI_{regional_clean}_{timestamp}.xlsx"

        return output_dir / filename

    def _create_dataframe(self, config: ExportConfig) -> pd.DataFrame:
        """Crea el DataFrame con las predicciones."""
        data_rows = []
        first_entry = next(iter(config.predictions_dict.values()))
        has_intervals = isinstance(first_entry, dict) and "limite_inferior" in first_entry

        for fecha_str, entry in config.predictions_dict.items():
            row = self._process_entry(
                fecha_str,
                entry,
                config.regional_nombre,
                include_ci=config.include_confidence_intervals,
                has_intervals=has_intervals,
            )
            data_rows.append(row)

        return pd.DataFrame(data_rows)

    def _process_entry(self, fecha_str: str, entry: Any, regional_nombre: str, *,
                      include_ci: bool, has_intervals: bool) -> dict[str, Any]:
        """Procesa una entrada de predicción individual."""
        if isinstance(entry, dict):
            return self._process_dict_entry(
                fecha_str,
                entry,
                regional_nombre,
                include_ci=include_ci,
                has_intervals=has_intervals,
            )

        return {
            "Fecha": fecha_str,
            f"SAIDI {regional_nombre} Predicho": round(float(entry), 2),
        }

    def _process_dict_entry(self, fecha_str: str, entry: dict[str, Any],
                       regional_nombre: str, *, include_ci: bool,
                       has_intervals: bool) -> dict[str, Any]:
        """Procesa entrada en formato diccionario."""
        valor = entry.get("valor_predicho", entry.get("valor", 0))
        row = {
            "Fecha": fecha_str,
            f"SAIDI {regional_nombre} Predicho": round(valor, 2),
        }

        if include_ci and has_intervals:
            self._add_confidence_intervals(row, entry)

        return row
    def _add_confidence_intervals(self, row: dict[str, Any], entry: dict[str, Any]):
        """Agrega intervalos de confianza al diccionario de fila."""
        inferior = entry.get("limite_inferior")
        superior = entry.get("limite_superior")
        margen = entry.get("margen_error")

        row["Limite Inferior (95%)"] = round(inferior, 2) if inferior is not None else None
        row["Limite Superior (95%)"] = round(superior, 2) if superior is not None else None
        row["Margen de Error"] = round(margen, 2) if margen is not None else None

    def _create_workbook(self, df: pd.DataFrame, config: ExportConfig) -> Workbook:  # noqa: ARG002
        """Crea el workbook de Excel con formato."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Predicciones SAIDI"

        self._write_headers(ws, df)
        self._write_data(ws, df)
        self._adjust_column_widths(ws)

        return wb

    def _write_headers(self, ws, df: pd.DataFrame):
        """Escribe los encabezados con formato."""
        styles = self._get_header_styles()

        for col_idx, column in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.fill = styles["fill"]
            cell.font = styles["font"]
            cell.alignment = styles["alignment"]
            cell.border = styles["border"]

    def _write_data(self, ws, df: pd.DataFrame):
        """Escribe los datos con formato."""
        cell_border = self._get_cell_border()
        data_alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, row_data in enumerate(df.values, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = data_alignment
                cell.border = cell_border

                if col_idx > 1 and value is not None:
                    cell.number_format = "0.00"

    def _adjust_column_widths(self, ws):
        """Ajusta el ancho de las columnas."""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except (ValueError, TypeError) as e:
                    print(f"Error al procesar celda: {e}")
                    continue

            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _add_model_info_sheet(self, wb: Workbook, config: ExportConfig):
        """Agrega hoja con información del modelo."""
        ws_info = wb.create_sheet("Informacion del Modelo")
        info_rows = self._build_model_info_rows(config)

        for row_idx, (label, value) in enumerate(info_rows, start=1):
            ws_info.cell(row=row_idx, column=1, value=label)
            ws_info.cell(row=row_idx, column=2, value=value)

            if label and not value:
                cell = ws_info.cell(row=row_idx, column=1)
                cell.font = Font(bold=True, size=11)

        ws_info.column_dimensions["A"].width = 30
        ws_info.column_dimensions["B"].width = 25

    def _build_model_info_rows(self, config: ExportConfig) -> list:
        """Construye las filas de información del modelo."""
        info_rows = [
            ["Informacion del Modelo de Prediccion", ""],
            ["", ""],
            ["Regional", config.regional_nombre],
            ["Codigo Regional", config.regional_code],
            ["Fecha de Exportacion", datetime.now(tz=UTC).strftime("%Y-%m-%d %H:%M:%S UTC")],
            ["", ""],
            ["Parametros del Modelo", ""],
        ]

        self._add_model_params(info_rows, config.model_info)
        self._add_model_metrics(info_rows, config.model_info)

        return info_rows

    def _add_model_params(self, info_rows: list, model_info: dict[str, Any]):
        """Agrega parámetros del modelo a las filas de información."""
        params_map = {
            "order": ("Order (p,d,q)", str),
            "seasonal_order": ("Seasonal Order (P,D,Q,s)", str),
            "transformation": ("Transformacion", lambda x: x.upper()),
            "with_exogenous": ("Variables Exogenas", lambda x: "SI" if x else "NO"),
            "with_simulation": ("Simulacion Climatica", lambda x: "SI" if x else "NO"),
            "confidence_level": ("Nivel de Confianza", lambda x: f"{x*100:.0f}%"),
        }

        for key, (label, formatter) in params_map.items():
            if key in model_info:
                info_rows.append([label, formatter(model_info[key])])

    def _add_model_metrics(self, info_rows: list, model_info: dict[str, Any]):
        """Agrega métricas del modelo a las filas de información."""
        if not model_info.get("metrics"):
            return

        info_rows.extend([["", ""], ["Metricas del Modelo", ""]])

        metrics = model_info["metrics"]
        metrics_map = {
            "rmse": ("RMSE", lambda x: f"{x:.4f}"),
            "mae": ("MAE", lambda x: f"{x:.4f}"),
            "mape": ("MAPE", lambda x: f"{x:.2f}%"),
            "r2_score": ("R2 Score", lambda x: f"{x:.4f}"),
            "precision_final": ("Precision Final", lambda x: f"{x:.2f}%"),
        }

        for key, (label, formatter) in metrics_map.items():
            if key in metrics:
                info_rows.append([label, formatter(metrics[key])])

    @staticmethod
    def _get_header_styles() -> dict[str, Any]:
        """Retorna los estilos para los encabezados."""
        return {
            "fill": PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
            "font": Font(bold=True, color="FFFFFF", size=12),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": ExportService._get_cell_border(),
        }

    @staticmethod
    def _get_cell_border() -> Border:
        """Retorna el borde para las celdas."""
        return Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def get_last_export_path(self):
        """Obtener la ruta del ultimo archivo exportado."""
        return self.last_export_path

    def export_to_custom_location(
        self,
        predictions_dict,
        regional_info: RegionalInfo,
        custom_path,
        *,
        include_confidence_intervals=True,
        model_info=None,
    ):
        """
        Exportar predicciones a una ubicacion especifica.

        Args:
            predictions_dict: Diccionario con predicciones
            regional_info: Informacion de la regional (codigo y nombre)
            custom_path: Ruta completa del archivo (incluyendo nombre)
            include_confidence_intervals: Incluir intervalos (keyword-only)
            model_info: Informacion del modelo (keyword-only)

        Returns:
            str: Ruta del archivo o None

        """
        if not predictions_dict:
            raise ValueError("No hay predicciones para exportar")

        try:
            custom_path_obj = Path(custom_path)

            # Si el archivo existe, eliminarlo primero para evitar conflictos
            if custom_path_obj.exists():
                try:
                    custom_path_obj.unlink()
                except (OSError, PermissionError) as e:
                    print(f"Advertencia: No se pudo eliminar archivo existente: {e}")

            # Crear configuración
            config = ExportConfig(
                predictions_dict=predictions_dict,
                regional_code=regional_info.code,
                regional_nombre=regional_info.nombre,
                output_dir=custom_path_obj.parent,
                include_confidence_intervals=include_confidence_intervals,
                model_info=model_info,
            )

            # Generar el archivo temporal con nombre automático
            temp_result = self._export_with_config(config)

            if not temp_result:
                return None

            # Renombrar al nombre deseado por el usuario
            temp_path = Path(temp_result)

            if temp_path.exists() and temp_path != custom_path_obj:
                try:
                    # Mover/renombrar al nombre final deseado
                    temp_path.rename(custom_path_obj)
                    self.last_export_path = str(custom_path_obj)
                    return str(custom_path_obj)
                except (OSError, PermissionError) as e:
                    print(f"Error al renombrar archivo: {e}")
                    # Si falla el renombre, al menos devolver la ruta temporal
                    return temp_result
            else:
                # Si ya tiene el nombre correcto o no existe temp
                return temp_result

        except (OSError, ValueError) as e:
            print(f"Error exportando a ubicacion personalizada: {e!s}")
            return None
